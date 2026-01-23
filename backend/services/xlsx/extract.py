from __future__ import annotations

import openpyxl

from backend.contracts import make_block
from backend.services.extract_utils import is_numeric_only, is_technical_terms_only

def extract_blocks(xlsx_path: str) -> dict:
    """
    Extract text blocks from an Excel file.
    
    Returns standard block format compatible with the translation pipeline.
    Uses read_only mode for performance with large files.
    """
    # Use read_only=True and data_only=True for speed and value extraction
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    blocks: list[dict] = []

    # Track metadata for the whole document
    for sheet_index, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]

        # Unique ID counter for blocks in this sheet
        block_id_counter = 0

        # Iterate through all cells that have values
        # Skills check: filter out numeric, technical terms, and keep structures
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                # Convert to string and clean
                text = str(cell.value).strip()

                # Use heuristics to filter non-translatable content
                if not text or is_numeric_only(text) or is_technical_terms_only(text):
                    continue

                block_id_counter += 1
                shape_id = block_id_counter

                # Standard block with extra Excel-specific fields
                block = make_block(
                    slide_index=sheet_index,
                    shape_id=shape_id,
                    block_type="spreadsheet_cell",
                    source_text=text
                )

                # Add Excel-specific metadata for reconstruction
                block["sheet_name"] = sheet_name
                block["cell_address"] = cell.coordinate

                # Optional: Capture style info if not in read_only mode,
                # but for bulk translation, consistency is key.

                blocks.append(block)

    return {
        "blocks": blocks,
        "sheet_count": len(wb.sheetnames)
    }

